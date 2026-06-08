"""
App Web — Bot DAS-SIMEI
Acesse em: http://localhost:5000
"""

import csv
import json
import os
import queue
import threading  # usado na thread do bot
from datetime import datetime
from io import StringIO
from pathlib import Path

from dotenv import load_dotenv, set_key
from flask import Flask, Response, jsonify, render_template, request, stream_with_context

import bot as bot_module

load_dotenv()

app = Flask(__name__)
app.secret_key = os.urandom(24)

CONFIG_FILE = Path("config.json")
UPLOAD_DIR  = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Estado global do processo
estado = {
    "rodando":   False,
    "parar":     [False],
    "log_queue": queue.Queue(),
}


# ── Config ────────────────────────────────────────────────────────────────────

def carregar_config() -> dict:
    if CONFIG_FILE.exists():
        return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    return {
        "pasta_downloads": str(Path("DAS").resolve()),
        "email_remetente": os.getenv("EMAIL_REMETENTE", ""),
        "email_senha":     os.getenv("EMAIL_SENHA_APP", ""),
        "email_destino":   os.getenv("EMAIL_DESTINATARIO", ""),
    }

def salvar_config(cfg: dict):
    CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    env_path = Path(".env")
    set_key(str(env_path), "EMAIL_REMETENTE",    cfg.get("email_remetente", ""))
    set_key(str(env_path), "EMAIL_SENHA_APP",    cfg.get("email_senha", ""))
    set_key(str(env_path), "EMAIL_DESTINATARIO", cfg.get("email_destino", ""))
    load_dotenv(override=True)


# ── Leitura de planilha ───────────────────────────────────────────────────────

def ler_planilha(caminho: Path) -> list[dict]:
    ext = caminho.suffix.lower()
    empresas = []

    if ext in (".xlsx", ".xlsm", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # Detecta linha de cabeçalho
        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        def col(row, *nomes):
            for n in nomes:
                for i, h in enumerate(header):
                    if n in h:
                        return str(row[i]).strip() if row[i] is not None else ""
            return ""
        for row in rows[1:]:
            if not any(row):
                continue
            empresas.append({
                "cnpj":  col(row, "cnpj"),
                "nome":  col(row, "nome", "razão", "empresa"),
                "email": col(row, "email", "e-mail"),
            })
        wb.close()

    elif ext == ".csv":
        text = caminho.read_text(encoding="utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        for row in reader:
            empresas.append({
                "cnpj":  row.get("cnpj", row.get("CNPJ", "")).strip(),
                "nome":  row.get("nome", row.get("Nome", row.get("NOME", ""))).strip(),
                "email": row.get("email", row.get("Email", row.get("EMAIL", ""))).strip(),
            })

    return [e for e in empresas if e["cnpj"]]


# ── Rotas ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/config", methods=["GET", "POST"])
def config():
    if request.method == "GET":
        return jsonify(carregar_config())
    cfg = request.json
    salvar_config(cfg)
    return jsonify({"ok": True})


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("planilha")
    if not f:
        return jsonify({"erro": "Nenhum arquivo enviado."}), 400
    ext = Path(f.filename).suffix.lower()
    if ext not in (".xlsx", ".xlsm", ".csv"):
        return jsonify({"erro": "Formato não suportado. Use .xlsx ou .csv"}), 400

    destino = UPLOAD_DIR / f"planilha{ext}"
    f.save(str(destino))

    try:
        empresas = ler_planilha(destino)
    except Exception as e:
        return jsonify({"erro": f"Erro ao ler planilha: {e}"}), 400

    return jsonify({"empresas": empresas, "total": len(empresas)})


@app.route("/iniciar", methods=["POST"])
def iniciar():
    if estado["rodando"]:
        return jsonify({"erro": "Processo já em andamento."}), 409

    dados = request.json or {}
    empresas = dados.get("empresas", [])
    cfg = carregar_config()
    pasta = Path(dados.get("pasta_downloads") or cfg["pasta_downloads"])

    if not empresas:
        return jsonify({"erro": "Nenhuma empresa para processar."}), 400

    # Limpa log anterior
    while not estado["log_queue"].empty():
        estado["log_queue"].get_nowait()

    estado["rodando"] = True
    estado["parar"] = [False]

    def callback_log(msg):
        estado["log_queue"].put(msg)

    def thread_bot():
        try:
            bot_module.executar(
                empresas=empresas,
                pasta_downloads=pasta,
                log_callback=callback_log,
                parar_flag=estado["parar"],
            )
        finally:
            estado["rodando"] = False
            estado["log_queue"].put("__FIM__")

    threading.Thread(target=thread_bot, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/parar", methods=["POST"])
def parar():
    estado["parar"][0] = True
    return jsonify({"ok": True})


@app.route("/status")
def status():
    return jsonify({"rodando": estado["rodando"]})




@app.route("/log-stream")
def log_stream():
    """Server-Sent Events — transmite o log em tempo real."""
    def gerar():
        while True:
            try:
                msg = estado["log_queue"].get(timeout=30)
                if msg == "__FIM__":
                    yield "data: __FIM__\n\n"
                    break
                yield f"data: {msg}\n\n"
            except queue.Empty:
                yield "data: __PING__\n\n"

    return Response(
        stream_with_context(gerar()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


if __name__ == "__main__":
    print("=" * 50)
    print("  Bot DAS-SIMEI — Interface Web")
    print("  Acesse: http://localhost:5000")
    print("=" * 50)
    app.run(host="0.0.0.0", debug=False, port=5000, threaded=True)
